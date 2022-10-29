#!/usr/bin/env python
# coding: utf-8

# In[3]:


import pandas as pd

from openpyxl import Workbook
import openpyxl
from openpyxl.styles import Font, Color

from openpyxl.chart import BarChart,Reference,Series,LineChart,ScatterChart

import xlsxwriter
import numpy as np
from sklearn.datasets import fetch_california_housing

import PyPDF2
import os
import re
import openpyxl

import docx2txt

import xlwt
import xml.etree.cElementTree as ET

user_input = int(input('''Enter 1 for Python program to Create a excel file
Enter 2 Python program for Import data from an excel file 
Enter 3 Python program for Format data in excel sheet 
Enter 4 Python program for Prepare bar chart
Enter 5 Python program for Prepare pie chart
Enter 6 Python program for Extract mobile no from PDF file and save into MS excel 
Enter 7 Python program for Extract mobile no from MS word file and save into MS excel
Enter 8 Python program for Extract mobile no from XML file and save into MS excel
''' ))

# 1) python program to create a excel file
if(user_input == 1):
    wb=Workbook()
    # print(wb.active.title)
    # print(wb.sheetnames)

    wb['Sheet'].title='Report of Automation'
    sh1=wb.active
    sh1['A1'].value='Date'
    sh1['B1'].value='Sales'
    sh1['A2'].value='01-11-21'
    sh1['B2'].value='1000'
    sh1['A3'].value='02-11-21'
    sh1['B3'].value='2000'
    sh1['A4'].value='03-11-21'
    sh1['B4'].value='2500'
    sh1['A5'].value='04-11-21'
    sh1['B5'].value='2000'
    sh1['A6'].value='05-11-21'
    sh1['B6'].value='3000'
    sh1['A7'].value='06-11-21'
    sh1['B7'].value='4000'
    sh1['A8'].value='07-11-21'
    sh1['B8'].value='3500'
    sh1['A9'].value='08-11-21'
    sh1['B9'].value='5000'
    sh1['A10'].value='09-11-21'
    sh1['B10'].value='7000'
    sh1['A11'].value='10-11-21'
    sh1['B11'].value='10000'
    wb.save("C:\\Users\\Murtuza pipulyawala\\Desktop\\sales.xlsx")
    
    
# 2) python program for import data from excel
if(user_input == 2):
    df = pd.read_csv("Yoshops_Order_List.csv",encoding="ISO-8859-1")
    df.head()
    
# 3) python program for format data in excel sheet 
if(user_input == 3):
    wb1 = openpyxl.load_workbook(r"C:\Users\Murtuza pipulyawala\Desktop\Yoshops_Order_List.xlsx")
    ws1 = wb1['Online Class']
    col_style = Font(name="Reem Kufi",color="DB3B22",underline='single')
    for i in range(2,101):
        ws1.cell(row=i,column=1).font = col_style

    wb1.save("C:\\Users\\Murtuza pipulyawala\\Desktop\\Yoshops_Order_List_formatted.xlsx")
    
# 4) python program for preparing bar chart
if(user_input == 4):
    wb=Workbook()
    ws=wb.active
    for i in range(10):
        ws.append([i])

    values = Reference(ws, min_col=1, min_row=1, max_col=1, max_row=10)
    chart = BarChart()
    ws.add_chart(chart,'A15')
    chart.title='Bar Chart'
    chart.y_axis.title='Test Number'
    chart.add_data(values)
    s1=chart.series[0]
    s1.marker.symbol = 'triangle'

    wb.save('C:\\Users\\Murtuza pipulyawala\\Desktop\\BarChart.xlsx')
    
# 5) python program for preparing pie chart
if(user_input == 5):
    cal = fetch_california_housing()

    cal_df = pd.DataFrame(data=np.c_[cal['data'], cal['target']],columns= list(cal['feature_names']) + ['target'])

    cal_df['AveRoomsRounded'] = cal_df['AveRooms'].round(decimals=0).astype(int)

    cal_df.loc[cal_df.AveRoomsRounded > 9, 'AveRoomsBin'] = '10+'
    cal_df.loc[(cal_df.AveRoomsRounded > 7) &  (cal_df.AveRoomsRounded < 10), 'AveRoomsBin'] = '8-9'
    cal_df.loc[(cal_df.AveRoomsRounded > 5) &  (cal_df.AveRoomsRounded < 8), 'AveRoomsBin'] = '6-7'
    cal_df.loc[(cal_df.AveRoomsRounded > 3) &  (cal_df.AveRoomsRounded < 6), 'AveRoomsBin'] = '4-5'
    cal_df.loc[cal_df.AveRoomsRounded < 4, 'AveRoomsBin'] = '<=3'


    cal_df_binned = cal_df['AveRoomsBin'].value_counts(normalize=True).rename_axis('AveRooms').reset_index(name='percentage')

    with pd.ExcelWriter('C:\\Users\\Murtuza pipulyawala\\Desktop\\california_housing_pie_chart.xlsx') as writer:
        cal_df_binned.to_excel(writer, index=False)
        wb = writer.book
        chart = wb.add_chart({'type': 'pie'})
        chart.set_title({'name': 'Average Rooms per Dwelling'})
        ws = writer.sheets['Sheet1']
        chart.add_series({'name': 'Average Rooms per Dwelling', 'values': '=Sheet1!$B$2:$B$6', 'categories': '=Sheet1!$A$2:$A$6'})
        ws.insert_chart('D2', chart)
        wb.save('C:\\Users\\Murtuza pipulyawala\\Desktop\\PieChart.xlsx')
        
# 6) python program for extract mobile no. from PDF
if(user_input == 6):
    load_pdf=open(r'C:\\Users\\Murtuza pipulyawala\\Desktop\\pdfdata.pdf','rb')

    read_pdf=PyPDF2.PdfFileReader(load_pdf)
    for i in range(0,6):
        first_page=read_pdf.getPage(i)
        page_content=first_page.extractText()
        page_content=page_content.replace('\n','')
        #print(page_content)
    #     mobile_number=re.search(r'(?:\+?\d{2}[-])?\d{10}',page_content)
    #     print(mobile_number)  
        pattern=re.compile(r'(?:\+?\d{2}[-])?\d{10}')
        phone_pdf=[]
        match=pattern.finditer(page_content)
        for x in match:
            phone_pdf.append(x.group(0))
    phone_pdf    
    from xlwt import Workbook
    wb1 = Workbook()
    sheet1 = wb1.add_sheet('sheet1')
    for i in range(len(phone_pdf)):
        sheet1.write(i,0,phone_pdf[i])
    wb1.save('C:\\Users\\Murtuza pipulyawala\\Desktop\\extract_phoneno_pdf.xls')    

# 7) python program for extract mobile no. from MS word
if(user_input == 7):
    read=docx2txt.process(r'C:\\Users\\Murtuza pipulyawala\\Downloads\\word data.docx')
    pattern=re.compile(r'\b[789]\d{9}\b')
    matches = pattern.finditer(read)
    phone_docx=[]
    for x in matches:
        phone_docx.append(x.group(0))  
    phone_docx  
    from xlwt import Workbook
    wb1 = Workbook()
    sheet1 = wb1.add_sheet('sheet1')
    for i in range(len(phone_docx)):
        sheet1.write(i,0,phone_docx[i])
    wb1.save('C:\\Users\\Murtuza pipulyawala\\Desktop\\extract_phoneno_docx.xls')    

# 8) python program for extract mobile no. from XML
if(user_input == 8):
    tree = ET.ElementTree(file='xmlDATA.xml')
    root = tree.getroot()
    xmlText = ''
    for page in root:
        for x in page:
            for y in x:
                if y.text != None:
                    xmlText+=y.text
    xmlText 
    pattern = re.compile(r'\b[789]\d{9}\b')
    matches = pattern.finditer(xmlText)
    phone_xml=[]
    for x in matches:
        phone_xml.append(x.group(0))
    phone_xml
#     from xlwt import Workbook
#     wb1 = Workbook()
#     sheet1 = wb1.add_sheet('sheet1')
#     for i in range(len(phone_xml)):
#         sheet1.write(i,0,phone_xml[i])
#     wb1.save('C:\\Users\\Murtuza pipulyawala\\Desktop\\extract_phoneno_xml.xls')    













